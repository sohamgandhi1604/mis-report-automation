import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DARK_BLUE  = "1F3864"
MID_BLUE   = "2F5496"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
GREY_ROW   = "F2F2F2"


def _apply_header_style(cell):
    cell.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=MID_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )


def _apply_body_style(cell, row_idx):
    bg = GREY_ROW if row_idx % 2 == 0 else WHITE
    cell.font      = Font(name="Calibri", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )


def _write_dataframe(ws, df, start_row=1, start_col=1):
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        _apply_header_style(cell)

    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=start_row + row_offset, column=col_idx, value=value)
            _apply_body_style(cell, row_offset)

    for col_idx, col_name in enumerate(df.columns, start=start_col):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df.iloc[:, col_idx - start_col])
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    return start_row + len(df) + 2


def _sheet_title(ws, title):
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value     = title
    cell.font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def _build_summary_sheet(wb, kpis, filename):
    ws = wb.active
    ws.title = "Executive Summary"
    _sheet_title(ws, "MIS Report — Executive Summary")

    ws["A3"] = "Report Generated:"
    ws["B3"] = datetime.now().strftime("%d %b %Y, %I:%M %p")
    ws["A4"] = "Source File:"
    ws["B4"] = os.path.basename(filename)
    ws["A3"].font = Font(name="Calibri", size=10, bold=True)
    ws["A4"].font = Font(name="Calibri", size=10, bold=True)

    ws["A6"] = "Key Performance Indicators"
    ws["A6"].font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)

    for i, (metric, value) in enumerate(kpis.items(), start=7):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        lc = ws.cell(row=i, column=1, value=metric)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font  = Font(name="Calibri", size=10, bold=True)
        vc.font  = Font(name="Calibri", size=10)
        lc.fill  = PatternFill("solid", fgColor=bg)
        vc.fill  = PatternFill("solid", fgColor=bg)
        side = Side(style="thin", color="BFBFBF")
        lc.border = Border(left=side, right=side, top=side, bottom=side)
        vc.border = Border(left=side, right=side, top=side, bottom=side)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22


def _build_cleaned_data_sheet(wb, df):
    ws = wb.create_sheet("Cleaned Data")
    _sheet_title(ws, "Cleaned & Standardised Data")
    display_df = df.head(10000)
    _write_dataframe(ws, display_df, start_row=3)
    ws.freeze_panes = "A4"


def _build_monthly_sheet(wb, monthly):
    if monthly.empty:
        return
    ws = wb.create_sheet("Monthly Revenue")
    _sheet_title(ws, "Monthly Revenue & Profit Trend")
    next_row = _write_dataframe(ws, monthly, start_row=3)

    rows_of_data = len(monthly) + 1
    chart = BarChart()
    chart.type         = "col"
    chart.title        = "Monthly Revenue"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Month"
    chart.style        = 10
    chart.width        = 22
    chart.height       = 13

    data = Reference(ws, min_col=2, min_row=3, max_row=3 + rows_of_data - 1)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + rows_of_data - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{next_row + 2}")


def _build_customers_sheet(wb, top_cust):
    if top_cust.empty:
        return
    ws = wb.create_sheet("Top Customers")
    _sheet_title(ws, "Top 10 Customers by Revenue")
    _write_dataframe(ws, top_cust.reset_index(drop=True), start_row=3)


def _build_products_sheet(wb, top_prod):
    if top_prod.empty:
        return
    ws = wb.create_sheet("Top Products")
    _sheet_title(ws, "Top 10 Products by Revenue")
    _write_dataframe(ws, top_prod.reset_index(drop=True), start_row=3)


def _build_region_sheet(wb, region):
    if region.empty:
        return
    ws = wb.create_sheet("Region Performance")
    _sheet_title(ws, "Region-Wise Performance")
    _write_dataframe(ws, region.reset_index(drop=True), start_row=3)


def generate_report(clean_df, kpis, monthly, top_customers,
                    top_products, region, source_path, output_path):
    wb = Workbook()
    _build_summary_sheet(wb, kpis, source_path)
    _build_cleaned_data_sheet(wb, clean_df)
    _build_monthly_sheet(wb, monthly)
    _build_customers_sheet(wb, top_customers)
    _build_products_sheet(wb, top_products)
    _build_region_sheet(wb, region)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"MIS_Report_{timestamp}.xlsx"
    full_path = os.path.join(output_path, filename)
    wb.save(full_path)
    return full_path